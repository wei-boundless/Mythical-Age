from __future__ import annotations

import csv
import json
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any

from capability_system.units.mcp.local.pdf.analysis.parser import PdfSegment, PdfTextParser

from .cleaner import ParsedContentCleaner
from .models import ParsedChunk


class MultimodalParserAdapter:
    """Local multimodal parser for the current agent stack.

    This adapter intentionally avoids hard dependency on RAG-Anything. Instead it
    parses common file types directly and emits unified chunks that can be passed
    to the existing vector indexer.
    """

    _SUPPORTED_EXTENSIONS = {
        ".pdf",
        ".png",
        ".jpg",
        ".jpeg",
        ".bmp",
        ".tiff",
        ".tif",
        ".gif",
        ".webp",
        ".doc",
        ".docx",
        ".ppt",
        ".pptx",
        ".xls",
        ".xlsx",
        ".txt",
        ".md",
        ".json",
        ".csv",
    }

    def __init__(
        self,
        repo_root: Path,
        *,
        ocr_language: str = "eng",
        max_pdf_pages: int = 30,
        max_csv_rows: int = 100,
        max_xlsx_rows_per_chunk: int = 20,
    ) -> None:
        self.repo_root = repo_root
        self.backend_root = (repo_root / "backend").resolve()
        self.ocr_language = ocr_language
        self.max_pdf_pages = max_pdf_pages
        self.max_csv_rows = max_csv_rows
        self.max_xlsx_rows_per_chunk = max(5, max_xlsx_rows_per_chunk)
        self.cleaner = ParsedContentCleaner()
        self._pdf_parser = PdfTextParser(root_dir=self.backend_root)

    def is_supported_file(self, path: Path) -> bool:
        return path.is_file() and path.suffix.lower() in self._SUPPORTED_EXTENSIONS

    def parser_available(self) -> bool:
        return True

    def capabilities(self) -> dict[str, bool]:
        return {
            "pdf_text": self._pdf_available(),
            "image_ocr": self._ocr_available(),
            "docx": True,
            "pptx": True,
            "xlsx": self._xlsx_available(),
        }

    def parse_file(self, path: Path) -> list[ParsedChunk]:
        suffix = path.suffix.lower()
        if suffix in {".txt", ".md"}:
            return self._parse_text_file(path)
        if suffix == ".json":
            return self._parse_json_file(path)
        if suffix == ".csv":
            return self._parse_csv_file(path)
        if suffix == ".pdf":
            return self._parse_pdf_file(path)
        if suffix in {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".gif", ".webp"}:
            return self._parse_image_file(path)
        if suffix == ".docx":
            return self._parse_docx_file(path)
        if suffix == ".pptx":
            return self._parse_pptx_file(path)
        if suffix == ".xlsx":
            return self._parse_xlsx_file(path)
        if suffix in {".doc", ".ppt", ".xls"}:
            return self._parse_binary_office_file(path)
        return []

    def _source(self, path: Path) -> str:
        resolved = path.resolve()
        try:
            relative = resolved.relative_to(self.backend_root)
        except ValueError as exc:
            raise ValueError(f"Refusing to parse file outside backend root: {resolved}") from exc
        return str(relative).replace(os.sep, "/")

    def _parse_text_file(self, path: Path) -> list[ParsedChunk]:
        text = self._clean_text(self._read_text_file(path), modality="text")
        if not text:
            return []
        return [
            ParsedChunk(
                text=text,
                source=self._source(path),
                modality="text",
                metadata={"parser": "native_multimodal"},
            )
        ]

    def _parse_json_file(self, path: Path) -> list[ParsedChunk]:
        raw_text = self._read_text_file(path)
        try:
            payload = json.loads(raw_text)
        except json.JSONDecodeError:
            return self._parse_text_file(path)

        lines: list[str] = []

        def walk(value: Any, prefix: str = "") -> None:
            if isinstance(value, dict):
                for key, sub_value in value.items():
                    next_prefix = f"{prefix}.{key}" if prefix else str(key)
                    walk(sub_value, next_prefix)
                return
            if isinstance(value, list):
                for idx, sub_value in enumerate(value):
                    next_prefix = f"{prefix}[{idx}]"
                    walk(sub_value, next_prefix)
                return
            text = str(value).strip()
            if text:
                lines.append(f"{prefix}: {text}" if prefix else text)

        walk(payload)
        joined = self._clean_text("\n".join(lines), modality="text")
        if not joined:
            return []
        return [
            ParsedChunk(
                text=joined,
                source=self._source(path),
                modality="text",
                metadata={"parser": "native_multimodal", "format": "json"},
            )
        ]

    def _parse_csv_file(self, path: Path) -> list[ParsedChunk]:
        rows = self._read_csv_rows(path, max_rows=self.max_csv_rows)
        rows = self.cleaner.clean_table_rows(rows)
        if len(rows) <= 1:
            return []

        chunks: list[ParsedChunk] = []
        for item in self._table_row_windows(
            sheet_name=path.stem,
            rows=rows,
            chunk_size=self.max_xlsx_rows_per_chunk,
            source_format="csv",
        ):
            chunks.append(
                ParsedChunk(
                    text=str(item["text"]),
                    source=self._source(path),
                    modality="table",
                    section=str(item["sheet_name"]),
                    metadata={
                        "parser": "native_multimodal",
                        "format": "csv",
                        "unit_view": "table_row_window",
                        "row_start": item.get("row_start"),
                        "row_end": item.get("row_end"),
                        "total_rows": item.get("total_rows"),
                        "chunk_index": item.get("chunk_index"),
                        "chunk_count": item.get("chunk_count"),
                        "header": item.get("header"),
                    },
                )
            )
        return chunks

    def _parse_pdf_file(self, path: Path) -> list[ParsedChunk]:
        if not self._pdf_parser.available():
            return []
        chunks: list[ParsedChunk] = []
        for segment in self._limit_pdf_segments(self._pdf_parser.extract_segments(path)):
            if not segment.text.strip():
                continue
            cleaned_text = self._clean_text(
                segment.text,
                modality=segment.modality,
                section=segment.section,
            )
            if not cleaned_text:
                continue
            metadata = {
                "parser": segment.metadata.get("parser", "pdf_parser"),
                "format": "pdf",
                **segment.metadata,
            }
            chunks.append(
                ParsedChunk(
                    text=cleaned_text,
                    source=self._source(path),
                    modality=segment.modality,
                    page=segment.page,
                    section=segment.section,
                    metadata=metadata,
                )
            )
        return chunks

    def _limit_pdf_segments(self, segments: list[PdfSegment]) -> list[PdfSegment]:
        if self.max_pdf_pages <= 0:
            return segments

        limited: list[PdfSegment] = []
        seen_pages: set[int] = set()
        for segment in segments:
            if segment.page is None:
                if len(seen_pages) < self.max_pdf_pages:
                    limited.append(segment)
                continue

            if segment.page not in seen_pages:
                if len(seen_pages) >= self.max_pdf_pages:
                    break
                seen_pages.add(segment.page)
            limited.append(segment)
        return limited

    def _parse_image_file(self, path: Path) -> list[ParsedChunk]:
        description_lines = [f"Image file: {path.name}"]
        metadata: dict[str, Any] = {"parser": "native_multimodal", "format": "image"}

        try:
            from PIL import Image  # type: ignore

            with Image.open(path) as image:
                metadata["width"], metadata["height"] = image.size
                metadata["mode"] = image.mode
                description_lines.append(
                    f"Image metadata: {image.size[0]}x{image.size[1]}, mode={image.mode}"
                )
        except Exception:
            pass

        ocr_text = self._extract_image_text(path)
        if ocr_text:
            description_lines.append("OCR text:")
            description_lines.append(ocr_text)
            metadata["ocr"] = True
        else:
            metadata["ocr"] = False

        cleaned_text = self._clean_text("\n".join(description_lines).strip(), modality="image")
        return [
            ParsedChunk(
                text=cleaned_text,
                source=self._source(path),
                modality="image",
                metadata=metadata,
            )
        ]

    def _parse_docx_file(self, path: Path) -> list[ParsedChunk]:
        texts = self._extract_docx_text(path)
        cleaned_text = self._clean_text("\n".join(texts), modality="text")
        if not cleaned_text:
            return []
        return [
            ParsedChunk(
                text=cleaned_text,
                source=self._source(path),
                modality="text",
                metadata={"parser": "native_multimodal", "format": "docx"},
            )
        ]

    def _parse_pptx_file(self, path: Path) -> list[ParsedChunk]:
        slides = self._extract_pptx_text(path)
        chunks: list[ParsedChunk] = []
        for slide_number, slide_text in slides:
            cleaned_slide = self._clean_text(slide_text, modality="text")
            if not cleaned_slide:
                continue
            chunks.append(
                ParsedChunk(
                    text=cleaned_slide,
                    source=self._source(path),
                    modality="text",
                    page=slide_number,
                    metadata={"parser": "native_multimodal", "format": "pptx"},
                )
            )
        return chunks

    def _parse_xlsx_file(self, path: Path) -> list[ParsedChunk]:
        sheets = self._extract_xlsx_chunks(path)
        chunks: list[ParsedChunk] = []
        for item in sheets:
            sheet_name = str(item["sheet_name"])
            sheet_text = str(item["text"])
            cleaned_sheet = self._clean_text(
                sheet_text,
                modality="table",
                section=sheet_name,
            )
            if not cleaned_sheet:
                continue
            metadata = {
                "parser": "native_multimodal",
                "format": "xlsx",
                "row_start": item.get("row_start"),
                "row_end": item.get("row_end"),
                "total_rows": item.get("total_rows"),
                "chunk_index": item.get("chunk_index"),
                "chunk_count": item.get("chunk_count"),
                "header": item.get("header"),
                "unit_view": "table_row_window",
            }
            chunks.append(
                ParsedChunk(
                    text=cleaned_sheet,
                    source=self._source(path),
                    modality="table",
                    section=sheet_name,
                    metadata=metadata,
                )
            )
        return chunks

    def _parse_binary_office_file(self, path: Path) -> list[ParsedChunk]:
        return [
            ParsedChunk(
                text=(
                    f"Binary Office file detected: {path.name}\n"
                    "Direct parsing is not implemented for this binary format. "
                    "Please convert it to docx/pptx/xlsx or PDF for better knowledge_system.retrieval."
                ),
                source=self._source(path),
                modality="text",
                metadata={"parser": "native_multimodal", "format": path.suffix.lower()},
            )
        ]

    def _read_text_file(self, path: Path) -> str:
        for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        return path.read_text(encoding="utf-8", errors="ignore")

    def _read_csv_rows(self, path: Path, *, max_rows: int) -> list[list[str]]:
        for encoding in ("utf-8", "gb18030", "latin-1"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.reader(handle)
                    rows: list[list[str]] = []
                    for idx, row in enumerate(reader):
                        if idx >= max_rows:
                            break
                        rows.append([cell.strip() for cell in row])
                    return rows
            except UnicodeDecodeError:
                continue
        return []

    def _extract_docx_text(self, path: Path) -> list[str]:
        paragraphs: list[str] = []
        try:
            with zipfile.ZipFile(path) as archive:
                xml_bytes = archive.read("word/document.xml")
        except Exception:
            return []

        root = ET.fromstring(xml_bytes)
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        for paragraph in root.findall(".//w:p", namespace):
            texts = [node.text for node in paragraph.findall(".//w:t", namespace) if node.text]
            merged = "".join(texts).strip()
            if merged:
                paragraphs.append(merged)
        return paragraphs

    def _extract_pptx_text(self, path: Path) -> list[tuple[int, str]]:
        slides: list[tuple[int, str]] = []
        try:
            with zipfile.ZipFile(path) as archive:
                slide_names = sorted(
                    name
                    for name in archive.namelist()
                    if name.startswith("ppt/slides/slide") and name.endswith(".xml")
                )
                namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
                for idx, slide_name in enumerate(slide_names, start=1):
                    root = ET.fromstring(archive.read(slide_name))
                    texts = [node.text for node in root.findall(".//a:t", namespace) if node.text]
                    merged = "\n".join(text.strip() for text in texts if text.strip()).strip()
                    if merged:
                        slides.append((idx, merged))
        except Exception:
            return []
        return slides

    def _extract_xlsx_chunks(self, path: Path) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook  # type: ignore

            workbook = load_workbook(filename=path, read_only=True, data_only=True)
            sheets: list[dict[str, Any]] = []
            for worksheet in workbook.worksheets:
                normalized_rows: list[list[str]] = []
                for row in worksheet.iter_rows(values_only=True):
                    cells = ["" if value is None else str(value).strip() for value in row]
                    while cells and not cells[-1]:
                        cells.pop()
                    if any(cells):
                        normalized_rows.append(cells)
                if len(normalized_rows) <= 1:
                    continue

                sheets.extend(
                    self._table_row_windows(
                        sheet_name=worksheet.title,
                        rows=normalized_rows,
                        chunk_size=self.max_xlsx_rows_per_chunk,
                        source_format="xlsx",
                    )
                )
            return sheets
        except Exception:
            return []

    def _table_row_windows(
        self,
        *,
        sheet_name: str,
        rows: list[list[str]],
        chunk_size: int,
        source_format: str,
    ) -> list[dict[str, Any]]:
        if len(rows) <= 1:
            return []
        header = rows[0]
        body_rows = rows[1:]
        total_rows = len(body_rows)
        window_size = max(1, int(chunk_size or 1))
        chunk_count = max(1, (total_rows + window_size - 1) // window_size)
        windows: list[dict[str, Any]] = []
        for offset in range(0, total_rows, window_size):
            part = body_rows[offset : offset + window_size]
            row_start = offset + 1
            row_end = offset + len(part)
            lines = [
                f"Table: {sheet_name}",
                f"Format: {source_format}",
                f"Columns: {' | '.join(header)}",
                f"Rows: {row_start}-{row_end} / {total_rows}",
            ]
            for row in part:
                normalized = row + [""] * (len(header) - len(row))
                lines.append(" | ".join(normalized[: len(header)]))
            windows.append(
                {
                    "sheet_name": sheet_name,
                    "text": "\n".join(lines),
                    "row_start": row_start,
                    "row_end": row_end,
                    "total_rows": total_rows,
                    "chunk_index": (offset // window_size) + 1,
                    "chunk_count": chunk_count,
                    "header": header,
                    "unit_view": "table_row_window",
                }
            )
        return windows

    def _extract_image_text(self, path: Path) -> str:
        if not self._ocr_available():
            return ""
        try:
            from PIL import Image  # type: ignore
            import pytesseract  # type: ignore

            tesseract_cmd = os.getenv("TESSERACT_CMD", "").strip()
            if tesseract_cmd:
                pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

            with Image.open(path) as image:
                text = pytesseract.image_to_string(image, lang=self.ocr_language)
        except Exception:
            return ""
        return self._normalize_ocr_text(text)

    def _clean_text(
        self,
        text: str,
        *,
        modality: str,
        section: str | None = None,
    ) -> str:
        return self.cleaner.clean_text(
            text,
            modality=modality,
            section=section,
        ).text

    def _normalize_ocr_text(self, text: str) -> str:
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        cleaned = "\n".join(lines)
        cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
        return cleaned.strip()

    def _pdf_available(self) -> bool:
        return self._pdf_parser.available()

    def _ocr_available(self) -> bool:
        try:
            import PIL  # type: ignore  # noqa: F401
            import pytesseract  # type: ignore  # noqa: F401
        except Exception:
            return False
        return True

    def _xlsx_available(self) -> bool:
        try:
            import openpyxl  # type: ignore  # noqa: F401
        except Exception:
            return False
        return True


